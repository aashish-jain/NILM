#!/usr/bin/env python
'''
    Non Intrusive device identification - characteristics extraction and identification
    Smart Devices group-Solarillion Foundation
'''

from openpyxl import (load_workbook,Workbook)
from statistics import mean
from math import sqrt


minimum_jump_magnitude=3
t=[]
devices=['D1','D2','D3','D4','D5']

#class for storing the parameters
class template_library():
    def __init__(self):
        self.first_maxima=0
        self.rate_of_change_transient=0
        self.settling_time=0
        self.avg_steadystate=0
        self.device=''
    def print_val(self):
        print '_'*80
        print self.device
        print "first_maxima={0} ".format(self.first_maxima)
        print " rate_of_change_transient={0} settling_time={1}ms".format(self.rate_of_change_transient,self.settling_time*20)
        print " avg_steadystate={0}".format(self.avg_steadystate)
        print '_'*80

# finds the settling instant in data list
def get_settling_instant(data,current_pos,start_pos):
    allowance=10
    a=0
    while(a!=3):
        temp=[]
        for i in xrange(current_pos,start_pos+75):
            temp.append(data[i])
        steady_state=mean(temp)
        for i in xrange(current_pos,start_pos+75):
            if (abs(data[i]-steady_state) <= 10):
                current_pos=i
                break
        del temp
        allowance-=5
        a+=1
    return current_pos

# extracts various identification parameters from Ipeak data
def extract_characteristics(data,device=""):
    t=template_library()                                                        #for storing the characteristics
    t.device=device
    previous_steadystate=data[0]                                                #Assuming first value of data to be SS value
    i=1                                                                         #setting index count of data to be 1
    #c1
    while(data[i]-previous_steadystate<minimum_jump_magnitude):                 #traverse the array till a jump is detected
        i+=1
    start_pos=i-1                                                               #instant before the device is switched on
    i+=1                                                                        #skipping the skew value
    #c2
    while(data[i]<data[i+1]):                                                   #finding first local maxima
        i+=1
    t.first_maxima=data[i]-previous_steadystate                                 #storing first maxima magnitude
    i+=1
    #transient and steady state average, settling time
    settling_instant = get_settling_instant(data,i,start_pos)
    t.settling_time = settling_instant - start_pos
    temp=[]
    while( i < settling_instant ):                                              #finding rate of change of Ipeak
        temp.append(data[i] - data[i+1])
        i+=1
    if(len(temp)!=0):
        t.rate_of_change_transient=mean(temp)
    del temp
    start_pos=i                                                                 #finding steady_state average
    while(i<len(data) and i-start_pos<10):
        t.avg_steadystate+=(data[i]-previous_steadystate);
        i+=1
    if(i != start_pos ):
        t.avg_steadystate/=(i-start_pos)
    return t                                                                    #returning the extracted characteristics

#reads data from given excel sheet (name of excel file and sheet should be same)
def read_excelsheet(device,c=0):
    wb = load_workbook('/home/aashish/DI/data/1/'+device+'.xlsx')               #create instance of load_workbook
    work_sheet=wb['Sheet1']                                                     #select the required worksheet
    c=randrange(65,85) if c==0 else c+64                                        #if no trial specified choose any random trail
    data=[]                                                                     #list to store Ipeak values
    for i in xrange(2,202):                                                      #copy values from cells and store it in data list
        data.append(work_sheet[chr(c)+str(i)].value)
    return extract_characteristics(data,device)                                 #return the extracted chracterisitcs from data

#takes the average values of exctracted characteristics from 20 trials
def avg(trials):
    temp=template_library()
    for i in xrange(0,len(trials)):
        temp.first_maxima += trials[i].first_maxima
        temp.rate_of_change_transient += trials[i].rate_of_change_transient
        temp.settling_time += trials[i].settling_time
        temp.avg_steadystate += trials[i].avg_steadystate
    temp.first_maxima /= 20
    temp.rate_of_change_transient /= 20
    temp.settling_time /= 20
    temp.avg_steadystate /= 20
    return temp

#witing the template in excel file
def create_library():
    print "Creating template library. Please wait..."
    wb=Workbook()
    ws=wb.active
    ws.title = 'Sheet1'

    row=67
    ws['B2']='Device'
    ws['B3']='first_maxima'
    ws['B4']='rate_of_change_transient'
    ws['B5']='settling_time'
    ws['B6']='avg_steadystate'

    for device in devices:
        trials=[]
        for trial_number in xrange(1,21):
            trials.append(read_excelsheet(device,trial_number))
        temp=(avg(trials))
        ws[chr(row)+'2']=device
        ws[chr(row)+'3']=temp.first_maxima
        ws[chr(row)+'4']=temp.rate_of_change_transient
        ws[chr(row)+'5']=temp.settling_time
        ws[chr(row)+'6']=temp.avg_steadystate
        row+=1

    wb.save('/home/aashish/DI/data/template.xlsx')
    print "creation successful !"

#Main code exectution starts here
create_library()
