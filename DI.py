#!/usr/bin/env python
'''
    Non Intrusive device identification - characteristics extraction and identification
    Smart Devices group-Solarillion Foundation
'''

from openpyxl import load_workbook
from random import randrange
from statistics import mean
from math import sqrt
'''
    PIP-> pyhton library manager
    to install library commmand is:
        pip install <Package_name>
    statistics.stdev library used for calculating standard devaition
    openpyxl.load_workbook for reading data from excel
    random.randrange for genrating random numbers within a range
'''


#constant(s) set after empirically observing the data
minimum_jump_magnitude=3
t=[]
devices=['D1','D2','D3','D4','D5']

class template_library():
    first_maxima=0
    avg_transient_ipeaks=0
    settling_time=0
    avg_steadystate=0
    device=''
    def print_val(self):
        print '_'*80
        print self.device
        print "first_maxima={0} ".format(self.first_maxima)
        print " avg_transient_ipeaks={0} settling_time={1} ms".format(self.avg_transient_ipeaks,self.settling_time*20)
        print " avg_steadystate={0}".format(self.avg_steadystate)
        print '_'*80

def get_settling_instant(data,current_pos,start_pos):
    temp=[]
    for i in xrange(current_pos,start_pos+75):
        temp.append(data[i])
    steady_state=mean(temp)
    for i in xrange(current_pos,start_pos+75):
        if (abs(data[i]-steady_state) <= 10):
            current_pos=i
            break
    del temp

    temp=[]
    for i in xrange(current_pos,start_pos+75):
        temp.append(data[i])
    steady_state=mean(temp)
    for i in xrange(current_pos,start_pos+75):
        if (abs(data[i]-steady_state) <= 5):
            current_pos=i
            break
    del temp

    temp=[]
    for i in xrange(current_pos,start_pos+75):
        temp.append(data[i])
    steady_state=mean(temp)
    for i in xrange(current_pos,start_pos+75):
        if (abs(data[i]-steady_state) <= 1):
            current_pos=i
            break
    del temp
    return current_pos

def extract_characteristics(data,device=""):                                    #for extracting characteristics
    t=template_library()                                                        #for storing the characteristics
    t.device=device
    previous_steadystate=data[0]                                                #Assuming first value of data to be SS value
    i=1                                                                         #setting index count of data to be 1
    #c1
    while(data[i]-previous_steadystate<minimum_jump_magnitude):                 #traverse the array till a jump is detected
        i+=1
    start_pos=i-1                                                               #instant before the device is switched on
    i+=1                                                                        #skipping the skew value
    #c2
    while(data[i]<data[i+1]):                                                   #finding first local maxima
        i+=1
    t.first_maxima=data[i]-previous_steadystate                                 #storing first maxima magnitude
    i+=1
    #transient and steady state average, settling time
    settling_instant = get_settling_instant(data,i,start_pos)
    print settling_instant
    t.settling_time = settling_instant - start_pos
    temp=[]
    while( i < settling_instant ):                                              #transient average calculation
        temp.append(data[i] - data[i+1])
        i+=1
    if(len(temp)!=0):
        # print temp
        t.avg_transient_ipeaks=mean(temp)
    del temp
    start_pos=i
    while(i<len(data) and i-start_pos<10):
        t.avg_steadystate+=(data[i]-previous_steadystate);
        i+=1
    if(i != start_pos ):
        t.avg_steadystate/=(i-start_pos)
    return t

#reads data from given excel sheet (name of excel file and sheet should be same)
def read_excelsheet(device,c=0):
    wb = load_workbook('/home/aashish/DI/data/1/'+device+'.xlsx')               #create instance of load_workbook
    work_sheet=wb['Sheet1']                                                     #select the required worksheet
    c=randrange(65,85) if c==0 else c+64                                        #if no trial specified choose any random trail
    data=[]                                                                     #list to store Ipeak values
    for i in xrange(2,202):                                                      #copy values from cells and store it in data list
        data.append(work_sheet[chr(c)+str(i)].value)
    return extract_characteristics(data,device)                                 #return the extracted chracterisitcs from data


def euclidean_distance_list(on_device):
    d=[]
    for i in xrange(0,len(t)):
        temp=[]
        temp.append((t[i].first_maxima - on_device.first_maxima)**2)
        temp.append((t[i].avg_transient_ipeaks - on_device.avg_transient_ipeaks)**2)
        temp.append((t[i].settling_time - on_device.settling_time)**2)
        temp.append((t[i].avg_steadystate - on_device.avg_steadystate)**2)
        # print i+1,' Temp= ',
        # for i in xrange(0,len(temp)):
        #     print (str(round(sqrt(temp[i]),2))).ljust(6),
        # print
        d.append(sqrt(abs(sum(temp))))
        del temp
    return d

def identify_device():
    for device in devices:
        print 'Actual Device : ',device,' '
        for i in xrange(1,21):
            distance=euclidean_distance_list(read_excelsheet(device,i))
            for i in xrange(0,len(devices)):
                print '_'*80
                print (str(round(distance[i],2))).ljust(6),
            min_index=0
            for i in xrange(1,len(distance)):
                if(distance[min_index]>distance[i]):
                    min_index=i
            print devices[min_index]

#Main code exectution starts here
trial_number=randrange(1,21)
print "trial_number ",trial_number
for device in devices:
    t.append(read_excelsheet(device,trial_number))
for i in xrange(0,len(t)):
    t[i].print_val()
# identify_device()
# (read_excelsheet('D4',trial_number)).print_val()
