#!/usr/bin/env python
'''
    Non Intrusive device identification - characteristics extraction and identification
    Smart Devices group-Solarillion Foundation
'''


from openpyxl import load_workbook
from random import randrange
from statistics import mean
from math import sqrt
'''
    PIP-> pyhton library manager
    to install library commmand is:
        pip install <Package_name>
    statistics.stdev library used for calculating standard devaition
    openpyxl.load_workbook for reading data from excel
    random.randrange for genrating random numbers within a range
'''

min_jump_magnitude=0
max_settling_time=0
t=[]
devices=['D1','D2','D3','D4','D5']


#class for storing the parameters
class template_library():
    def __init__(self):
        self.first_maxima=0
        self.rate_of_change_transient=0
        self.settling_time=0
        self.avg_steadystate=0
        self.device=''
    def print_val(self):
        print '_'*80
        print self.device
        print "first_maxima={0} ".format(self.first_maxima)
        print " rate_of_change_transient={0} settling_time={1}ms".format(self.rate_of_change_transient,self.settling_time*20)
        print " avg_steadystate={0}".format(self.avg_steadystate)
        print '_'*80

# finds the settling instant in data list
def get_settling_instant(data,current_pos,start_pos):
    allowance=11
    a=0
    while(allowance!=0):
        temp=[]
        for i in xrange(current_pos,start_pos+max_settling_time*2):
            if(i == len(data)):
                break
            temp.append(data[i])
        steady_state=mean(temp)
        for i in xrange(current_pos,max_settling_time*2):
            if (abs(data[i]-steady_state) <= allowance):
                current_pos=i
                break
        del temp
        allowance-=1
    return current_pos

# extracts various identification parameters from Ipeak data
def extract_characteristics(data,device=""):
    T=[]
    start_pos=0
    while(1):
        t=template_library()                                                    #for storing the characteristics
        t.device=device
        previous_steadystate=data[start_pos]                                    #Assuming first value of data to be SS value
        i=start_pos+1                                                           #setting index count of data to be 1
        #traverse the array till a jump is detected
        while(data[i]-previous_steadystate<minimum_jump_magnitude and i<len(data)-1):
            i+=1
        if(i==len(data)-1):                                                     #end of the list encountered
            break
        start_pos=i-1                                                           #instant before the device is switched on
        i+=1                                                                    #skipping the skew value
        # first local maxima
        while(data[i]<data[i+1]):                                               #finding first local maxima
            i+=1
        t.first_maxima=data[i]-previous_steadystate                             #storing first maxima magnitude
        i+=1
        #transient and steady state average, settling time
        settling_instant = get_settling_instant(data,i,start_pos)
        t.settling_time = settling_instant - start_pos
        temp=[]
        while( i < settling_instant ):                                          #finding rate of change of Ipeak
            temp.append(data[i] - data[i+1])
            i+=1
        if(len(temp)!=0):
            t.rate_of_change_transient=mean(temp)
        del temp
        start_pos=i                                                             #finding steady_state average
        while(i<len(data) and i-start_pos<10):
            t.avg_steadystate+=(data[i]-previous_steadystate);
            i+=1
        if(i != start_pos ):
            t.avg_steadystate/=(i-start_pos)
        start_pos=i
        T.append(t)
        if(len(data)-i <max_settling_time*2):
            break
        # print 'hi'
        # raw_input()
        # for x in xrange(0,len(T)):
        #     T[x].print_val()
    return T                                                                    #returning the extracted characteristics

#reads data from given excel sheet (name of excel file and sheet should be same)
def read_excelsheet(device,c=0):
    wb = load_workbook('/home/aashish/DI/find/'+device+'.xlsx')                 #create instance of load_workbook
    work_sheet=wb['Sheet1']                                                     #select the required worksheet
    c=randrange(65,85) if c==0 else c+64                                        #if no trial specified choose any random trail
    data=[]                                                                     #list to store Ipeak values
    i=1
    value=0
    while(1):                                                                    #copy values from cells and store it in data list
        value=work_sheet[chr(c)+str(i)].value
        if(value==None):
            break
        data.append(work_sheet[chr(c)+str(i)].value)
        i+=1
    return extract_characteristics(data,device)                                 #return the extracted chracterisitcs from data

#reads the created device paramter template
def read_template():
    global max_settling_time
    global min_jump_magnitude
    wb = load_workbook('/home/aashish/DI/data/template.xlsx')
    ws = wb['Sheet1']
    min_jump_magnitude=ws['C1'].value
    max_settling_time=ws['F1'].value
    row=67
    for row in range(67,72):
        temp=template_library()
        temp.device=ws[chr(row)+'2'].value
        temp.first_maxima=ws[chr(row)+'3'].value
        temp.rate_of_change_transient=ws[chr(row)+'4'].value
        temp.settling_time=ws[chr(row)+'5'].value
        temp.avg_steadystate=ws[chr(row)+'6'].value
        t.append(temp)
        row+=1

# calulates distance between the device and tempalte device points
def euclidean_distance_list(on_device):
    d=[]
    for i in xrange(0,len(t)):
        for j in xrange(0,len(t)):
            temp=[]
        temp.append((t[i].first_maxima - on_device.first_maxima)**2)
        temp.append((t[i].rate_of_change_transient - on_device.rate_of_change_transient)**2)
        temp.append((t[i].settling_time - on_device.settling_time)**2)
        temp.append((t[i].avg_steadystate - on_device.avg_steadystate)**2)
        # print i+1,' Temp= ',
        # for i in xrange(0,len(temp)):
        #     print (str(round(sqrt(temp[i]),2))).ljust(6),
        # print
        d.append(sqrt(abs(sum(temp))))
        del temp
    return d

def identify_device():
    for device in devices[5:]:
        print 'Actual Device : ',device,' '
        on_device=[]
        for trial in xrange(1,21):
            on_device.append(read_excelsheet(device,trial))
        for j in xrange(0,len(on_device)):
            for k in xrange(0,len(on_device[j])):
                distance=euclidean_distance_list(on_device[j][k])
                # for i in xrange(0,len(devices)):
                #     print '_'*80
                #     print (str(round(distance[i],2))).ljust(6),
                min_index=0
                for i in xrange(1,len(distance)):
                    if(distance[min_index]>distance[i]):
                        min_index=i
                print devices[min_index],' ',
            print
        del on_device

#Main code exectution starts here
devices.append(raw_input("Enter device name"))
read_template()
for i in xrange(0,len(t)):
    t[i].print_val()
identify_device()
